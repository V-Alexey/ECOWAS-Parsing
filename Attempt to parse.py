#!/usr/bin/env python
# coding: utf-8

# # Functions

# In[4]:


import re                # Module To work with strings later

# To unite words (and roots of these words) that we are interested in in one dictionary 
def sift_words (test_d):          # Among all the mentioned words we take those that we are interested in
    tag_list = ['TERROR', 'EXTREMIS', 'SECUR', 'ECONOM', 'TRAD', 'MARKET'] # Roots of the words
    save_dict = {}
    Dictionary = {'TERRORISM': 0, 'EXTREMISM': 0, 'SECURITY': 0, 'ECONOMIC': 0, 'TRADE': 0, 'MARKET': 0} # The words themselves
    for i in tag_list:
        for el in test_d.items():
            if el[0].startswith(i): # Assign words instead of tags, so we can count them
                save_dict[el[0]] = el[1]
    
    for element in save_dict.items():   # Counting the number of mentions of the certain words
        if element[0].startswith('TERROR'):
            Dictionary['TERRORISM'] += element[1]
        elif element[0].startswith('EXTREMIS'):
            Dictionary['EXTREMISM'] += element[1]
        elif element[0].startswith('SECUR'):
            Dictionary['SECURITY'] += element[1]
        elif element[0].startswith('ECONOM'):
            Dictionary['ECONOMIC'] += element[1]
        elif element[0].startswith('TRAD'):
            Dictionary['TRADE'] += element[1]
        elif element[0].startswith('MARKET'):
            Dictionary['MARKET'] += element[1]
    return(Dictionary)                   # Saving everything into the dictionary


# In[5]:


# To sum up all the mentions throughout years (for personal stats, not for use) => getting picture of what is popular overall

def dic_append (sifted, Dictionary):   
    for element in sifted.items():
        if element[0].startswith('TERROR'):
            Dictionary['TERRORISM'] += element[1]
        elif element[0].startswith('EXTREMIS'):
            Dictionary['EXTREMISM'] += element[1]
        elif element[0].startswith('SECUR'):
            Dictionary['SECURITY'] += element[1]
        elif element[0].startswith('ECONOM'):
            Dictionary['ECONOMIC'] += element[1]
        elif element[0].startswith('TRAD'):
            Dictionary['TRADE'] += element[1]
        elif element[0].startswith('MARKET'):
            Dictionary['MARKET'] += element[1]

    return(Dict)


# In[6]:


import textwrap # to remove extra indents

# Getting rid of <span>

def clear_date (Date):
    
    Not_wanted = ['<', '>', '/', '\n'] # List of what we don't want to see
    Finish = []
    goal = Date.prettify() # to divide among the tags and make it smoother

    for i in goal:
        if i not in Not_wanted:
            Finish.append(i)      # Throwing out all the not wanted parts
            new_finish = ''.join(Finish) # Joinging into the final string

    Fin = textwrap.dedent (new_finish.replace("span", "")) # Getting rid of span

    # Declare what Month corresponds with what number
    Seasons = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,'Sep': 9, 'Oct': 10, 
               'Nov': 11, 'Dec': 12}


#Making Date the conventional type (/ or .)
    Integers = []
    Words = []
    Year = []
    for element in Fin:
        if element.isdigit(): # getting all the digits in here (tryong to separate digits, from the short name of month)
            Integers.append(element)


    Year = Integers[-4:] # Getting the last 4 digits, which always will be year for this web-site
    del Integers [-4:] # Clearing the lsat 4 elements of the list for future use
    Integers = int(''.join(Integers)) # making the remaining numbers an integers (these will be days) 
    Year = int(''.join(Year)) # Making years (last 4 digits) integers

    for letter in Fin:
        if letter.isalpha(): # If it is not a digit (month)
            Words.append(letter)
            word = ''.join(Words) # We make it into a variable
            new_word = textwrap.dedent(word.replace (',', '')) # Getting rid of all the indents

    for a, bs in Seasons.items(): # Converting Seasons into numbers
        if new_word in Seasons:
            Month = (Seasons[new_word]) # If a variable (line 41) is in declared dictionary, then give us the corresponding number

    numbers = [Integers, Month, Year] # Unite everything in one list


    result = '/'.join(map(str, numbers)) # Separate them with /
    return(result, Year)


# # Parsing (around 2-2.5 hours)

# In[7]:


from bs4 import BeautifulSoup # Module For parsing
import requests               # To read HTML links easily
import xlsxwriter             # To create and append excel file with data
from collections import Counter # To count words in the news 

Dict = {'TERRORISM': 0, 'EXTREMISM': 0, 'SECURITY': 0, 'ECONOMIC': 0, 'TRADE': 0, 'MARKET': 0} # Dicitonary that we are interested in
base_URL = 'https://www.ecowas.int/category/news/' # The main page of news of the web-site
current_URL = None    # For now the URL is noen, so we can change it

work_name = 'course_paper8.xlsx'    # Declaring the excel file name
workbook = xlsxwriter.Workbook(work_name) # Create excel file
worksheet = workbook.add_worksheet('Primary Data') # Name sheet in excel file
headers = ['Date', 'Year','TERRORISM', 'EXTREMISM', 'SECURITY', 'ECONOMIC', 'TRADE', 'MARKET'] 
worksheet.write_row(0, 0, headers) # Fill in the first row of excel file with the declared headers
col = 2
row = 1

for page in range(1,74): # creating loop, so we can change pages (all in all there are 74 pages of news)
    current_URL = base_URL + f"page/{page}/" # the pages change only with the number? which is convenient
    r = requests.get(current_URL)            # Getting the URL 
    soup = BeautifulSoup(r.content, 'html.parser') # Getting the content of the URL and reading it
    for h3 in soup.findAll("h3"): # h3 on a web-site is a heading with the linnk attached => looking for all the geadings
        for a in h3.findAll("a", {"href": True}): # Finding all the links attached to the heading
            new_URL = a["href"] 
            n_r = requests.get(new_URL)
            new_soup = BeautifulSoup(n_r.content, 'html.parser') # Going to the link of the heading of a news
            div_text = new_soup.find("div", {"class": "article-content"}).get_text() # looking at what is written in the news => getting text from there
            Date = new_soup.find("div", class_="article-content").find("span") # The Date (made very bad on the site) is stored in the <span> tag, so we need to take it
            Fin_date = clear_date(Date) # Unsing the earlier created function to get the cleared Date

            worksheet.write(row, col-2, Fin_date[0]) # We write in first column the full Date
            worksheet.write(row, col-1, Fin_date[1]) # And in the Second column only the year for working with it
            #We have text from the website (one piece of news)

            # Counting words that appear in the text
            interm = div_text.replace("\n", "") # GEtting rid of paragraphs
            interm1 = interm.replace("\xa0", "") # and whatever this is (appears in the text very often => poor design of the site)
            fin_text = interm1.upper().split(' ')
            All_Mentions_per_news = dict(Counter(fin_text)) # This is all the mentions for a piece of news without sorting words
            Mentions_per_news = sift_words(All_Mentions_per_news) # This is sorted (sifted) => show only tags that we are interested in
            All_mentions = dic_append(Mentions_per_news, Dict) # This gves us sifted tags, but all the mentions in all years


            for key, value in Mentions_per_news.items():
                worksheet.write(row, col, value) # Writing in excel file the info about the 6 collected tags
                col += 1
            row += 1        #changing rows and columns, so it could loop properly again
            col = 2

            print(Fin_date) # Printing here is just for making sure that everything is working (does not influence the code)


workbook.close()  #Closing the workbook after every piece of news was parsed


# # Counting the results by years to get into a table 

# In[93]:


from openpyxl import load_workbook # Module allows to work with existing workbooks and append necessary information there
import pandas as pd                # Pandas to get access to created dataframe

df = pd.read_excel(work_name) # Reading created file with pandas
wb = load_workbook(filename = work_name) # Opening the created file in python
ws = wb.active
headers = ['Terrorism', 'Extremism', 'Security', 'Economic', 'Trade', 'Market']
row = 2
col = 10 # Choosing the place of the future table
Ind = 0

for i in range (6): # Creating column of the tags that we have collected
    Cell = ws.cell(row = row, column = col).value = headers[Ind]
    row += 1
    Ind += 1

Year = 2017
row = 1
col = 11
for i in range (7): # Creating the row of years that we hae information for
    Cell = ws.cell(row = row, column = col).value = Year
    col += 1
    Year += 1

row = 2
col = 11
Year = 2017
while Year != 2024: # The timeframe of the paper - 2023, so we don't need 2024.
    row = 2
    variable = 'y' + f"_{Year}" # We create variable that cooresponds to the year
    variable = df[df['Year'] == Year] # assigning to it the sorted by year (in pandas) dataframe
    Cell = ws.cell(row = row, column = col).value = variable['TERRORISM'].sum() # Adding to the cell in excel the sum for a tag in a certain year
    row += 1  #Shifting the row to the one below
    Cell = ws.cell(row = row, column = col).value = variable['EXTREMISM'].sum() # Repeating with another tag
    row += 1
    Cell = ws.cell(row = row, column = col).value = variable['SECURITY'].sum()
    row += 1
    Cell = ws.cell(row = row, column = col).value = variable['ECONOMIC'].sum()
    row += 1
    Cell = ws.cell(row = row, column = col).value = variable['TRADE'].sum()
    row += 1
    Cell = ws.cell(row = row, column = col).value = variable['MARKET'].sum()
    Year += 1 # Adding + 1 year, so we have new information
    col += 1  #Shofting one column right
    
wb.save(work_name) # When the program is done, it rewrites the file, so we don't consume much space



